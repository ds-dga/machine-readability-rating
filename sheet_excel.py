from openpyxl import load_workbook
import pandas as pd
from itertools import chain

from sheet_common import has_na_in_number_cols

""" XLS, XLSX grader
1. check if it's excel file.
2. check if it has one header row. This is a bit triggy how to find that out.
    - has header or not
    - has one-row header or multiple row
3. data consistency
"""


def has_good_header(fpath):
    wb = load_workbook(filename=fpath)
    sheets = wb.sheetnames
    for sn in sheets:
        ws = wb[sn]
        cell = ws["A1"].value
        if not cell:
            return False, "empty A1"
        for cell in list(ws.rows)[0]:
            cv = cell.value
            if not cv or cv in ".-0123456789":
                return False, "header not alpha"
    return True, ""


def is_shape_consistency(fpath, **kwargs):
    """check dimension and especially column consistency

    return when found inconsistensy as fast as possible
        True        consistency
        False       inconsistency
    """
    wb = load_workbook(
        filename=fpath,
        data_only=True,
        read_only=True,
    )
    shapes = []
    sheets = wb.sheetnames
    for sn in sheets:
        try:
            ws = wb[sn]
            rows = list(ws.rows)
            shapes.append([len(rows) - 1, len(rows[0])])
        except IndexError:
            pass
        # openpyxl doesn't handle 1-row as header automaticall while
        # pandas always assumes that -- we need to make it fair for both cases

    df_shapes = []
    try:
        dfs = pd.read_excel(fpath, sheet_name=None)
        for sheet in dfs.keys():
            df_shapes.append(dfs[sheet].shape)
    except pd.errors.DtypeWarning:
        return False, "df: mixed types"

    good = list(chain(*shapes)) == list(chain(*df_shapes))
    # if not good:
    #     print(shapes, df_shapes)
    #     print([c.value for c in rows[0]])
    #     print(dfs[sheets[len(sheets) - 1]].iloc[0])
    return good, "" if good else "pyxl != dataframe"


def has_merged_cells(fpath, **kwargs):
    """check if there is any merged cell

    return when found inconsistensy as fast as possible
        True        found merged cells & sheet + cound
        False       not found, ''
    """
    wb = load_workbook(filename=fpath)
    sheets = wb.sheetnames
    for sn in sheets:
        try:
            ws = wb[sn]
            merged_cells = ws.merged_cells.ranges
            if not merged_cells:
                continue
            bd = [f"{i}" for i in list(merged_cells[0].bounds)]
            return True, f"{sn}/{','.join(bd)}"
        except IndexError:
            pass
    return False, ""


def are_num_cols_consistency(fpath):
    """Basically check if there is NA or null in numbered columns

    return <bool>
        True    all numbered columns are consistency
        False   Not all numbered columns are consistency
    """
    df = pd.read_excel(fpath)
    bad, _ = has_na_in_number_cols(df)
    return not bad
